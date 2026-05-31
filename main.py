"""
영화 추천 서비스 - FastAPI 백엔드
======================================
동작 모드 2가지:
  1. DB 연결 성공  → MySQL에서 데이터 조회 (운영 환경)
  2. DB 연결 실패  → 엑셀 파일에서 데이터 로드 (개발/테스트 환경)

팀원 연동 안내:
  - DB 담당자: mysql 컨테이너 준비 후 환경변수 설정하면 자동으로 DB 모드로 전환
  - 프론트엔드: /api/movies?emotion=울고+싶어  형태로 호출
"""

import os
import re
from contextlib import asynccontextmanager
from typing import Optional

import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

# ── 환경 변수 (docker-compose에서 주입, 없으면 기본값) ─────────────────────────
DB_HOST     = os.getenv("MYSQL_HOST", "")          # 비어있으면 엑셀 모드
DB_USER     = os.getenv("MYSQL_USER", "root")
DB_PASSWORD = os.getenv("MYSQL_PASSWORD", "root1234")
DB_NAME     = os.getenv("MYSQL_DB", "movie_recommend_db")
EXCEL_PATH  = os.getenv("EXCEL_PATH", "data/movies.xlsx")

# ── 전역 데이터 저장소 (엑셀 모드용) ──────────────────────────────────────────
_movies: dict[str, dict] = {}       # movie_id → movie dict
_ott:    dict[str, list] = {}       # movie_id → [ott dict, ...]
_use_db = False
_db_pool = None


# ── 라벨 정규화 (괄호 제거, 오탈자 보정) ──────────────────────────────────────
LABEL_FIXES = {
    "미장센이 좋은 영화를 좋은 영화를 보고 싶어": "미장센이 좋은 영화를 보고 싶어",
}

def _normalize_label(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    label = raw.strip().strip("()")
    return LABEL_FIXES.get(label, label)


# ── 엑셀 로더 ─────────────────────────────────────────────────────────────────
def load_excel(path: str):
    """엑셀 파일을 읽어 _movies, _ott 전역 딕셔너리를 채웁니다."""
    global _movies, _ott

    if not os.path.exists(path):
        print(f"[WARN] 엑셀 파일 없음: {path}  →  data/ 폴더에 movies.xlsx를 넣어주세요.")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── Movies 시트 ────────────────────────────────────────────────────────────
    ws_movies = wb["Movies"]
    for i, row in enumerate(ws_movies.iter_rows(values_only=True)):
        if i == 0:  # 헤더 스킵
            continue
        movie_id = row[0]
        if not movie_id:
            continue

        rating_raw = row[5]
        try:
            rating = round(float(rating_raw), 1) if rating_raw else None
        except (ValueError, TypeError):
            rating = None

        _movies[str(movie_id)] = {
            "id":         str(movie_id),
            "title":      row[1] or "",
            "genre":      row[2] or "",
            "year":       int(row[3]) if row[3] else None,
            "time":       f"{row[4]}분" if row[4] else None,
            "rating":     rating,
            "label1":     _normalize_label(row[6]),
            "label2":     _normalize_label(row[7]),
            "is_certain": int(row[8]) if row[8] is not None else 0,
            "pos":        row[9] or "",
            "neg":        row[10] or "",
            "ott":        [],   # OTT 시트 처리 후 채워짐
        }

    # ── OTT_Offers 시트 ────────────────────────────────────────────────────────
    ws_ott = wb["OTT_Offers"]
    for i, row in enumerate(ws_ott.iter_rows(values_only=True)):
        if i == 0:
            continue
        movie_id, platform, offer_type, rent_price, buy_price = (
            row[0], row[1], row[2], row[3], row[4]
        )
        if not movie_id or not platform:
            continue
        # 엑셀 오타 보정: 'M42' → 'M042' (M + 2자리 숫자인 경우 0 패딩)
        raw = str(movie_id).strip()
        import re as _re
        mid = _re.sub(r'^(M)(\d{2})$', lambda m: m.group(1) + '0' + m.group(2), raw)
        if mid not in _ott:
            _ott[mid] = []
        _ott[mid].append({
            "platform":   platform,
            "type":       offer_type,
            "rent_price": int(rent_price) if rent_price else None,
            "buy_price":  int(buy_price)  if buy_price  else None,
        })

    # OTT를 영화 딕셔너리에 합치기
    for mid, movie in _movies.items():
        movie["ott"] = _ott.get(mid, [])

    wb.close()
    print(f"[INFO] 엑셀 로드 완료: {len(_movies)}편 / OTT {sum(len(v) for v in _ott.values())}건")


# ── DB 연결 시도 ───────────────────────────────────────────────────────────────
def try_connect_db() -> bool:
    """DB 연결 가능 여부 반환. 가능하면 _db_pool 설정."""
    global _use_db, _db_pool
    if not DB_HOST:
        print("[INFO] MYSQL_HOST 미설정 → 엑셀 모드로 실행")
        return False
    try:
        import mysql.connector
        _db_pool = mysql.connector.connect(
            host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME
        )
        _use_db = True
        print(f"[INFO] MySQL 연결 성공: {DB_HOST}/{DB_NAME}")
        return True
    except Exception as e:
        print(f"[WARN] MySQL 연결 실패 ({e}) → 엑셀 모드로 실행")
        return False


# ── FastAPI 앱 ─────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    if not try_connect_db():
        load_excel(EXCEL_PATH)
    yield
    if _db_pool:
        try:
            _db_pool.close()
        except Exception:
            pass


app = FastAPI(title="영화 추천 API", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── 헬퍼: 영화 목록 가져오기 ──────────────────────────────────────────────────
def _get_movies_by_emotion(emotion: str, order_by: str = "rating") -> list[dict]:
    """emotion 문자열과 label1/label2 매칭 → 영화 목록 반환."""
    if _use_db:
        return _db_get_movies_by_emotion(emotion, order_by)
    return _excel_get_movies_by_emotion(emotion, order_by)


def _excel_get_movies_by_emotion(emotion: str, order_by: str = "rating") -> list[dict]:
    result = []
    for m in _movies.values():
        if emotion in (m.get("label1") or "") or emotion in (m.get("label2") or ""):
            result.append(m)
            
    # 프론트가 준 order_by 값에 따라 다르게 정렬
    if order_by == "year":
        # 최신순 (개봉 연도 큰 순서대로)
        result.sort(key=lambda x: (-(x["year"] or 0)))
    else:
        # 별점순 (기존 유지)
        result.sort(key=lambda x: (-(x["is_certain"] or 0), -(x["rating"] or 0)))
        
    return result


def _db_get_movies_by_emotion(emotion: str, order_by: str = "rating") -> list[dict]:
    cur = _db_pool.cursor(dictionary=True)
    
    # 기본은 별점순 정렬 SQL
    order_sql = "ORDER BY m.is_certain DESC, m.rating DESC"
    
    # 만약 프론트에서 최신순(year)을 요청하면 SQL 정렬 조건을 연도순으로 변경
    if order_by == "year":
        order_sql = "ORDER BY m.year DESC"
        
    query = f"""
        SELECT m.movie_id AS id, m.title, m.genre, m.year,
               CONCAT(m.runtime, '분') AS time,
               m.rating, m.label1, m.label2, m.is_certain,
               m.positive_review AS pos, m.negative_review AS neg
        FROM movies m
        WHERE m.label1 = %s OR m.label2 = %s
        {order_sql}
    """
    
    cur.execute(query, (emotion, emotion))
    rows = cur.fetchall()
    cur.close()

    for row in rows:
        row["ott"] = _db_get_ott(row["id"])
        if row["rating"]:
            row["rating"] = float(row["rating"])
    return rows


def _db_get_ott(movie_id: str) -> list[dict]:
    cur = _db_pool.cursor(dictionary=True)
    cur.execute(
        "SELECT platform, type, rent_price, buy_price FROM ott_offers WHERE movie_id = %s",
        (movie_id,),
    )
    rows = cur.fetchall()
    cur.close()
    return rows


def _get_movie_detail(movie_id: str) -> Optional[dict]:
    if _use_db:
        return _db_get_movie_detail(movie_id)
    return _movies.get(movie_id)


def _db_get_movie_detail(movie_id: str) -> Optional[dict]:
    cur = _db_pool.cursor(dictionary=True)
    cur.execute(
        """
        SELECT movie_id AS id, title, genre, year,
               CONCAT(runtime, '분') AS time, rating,
               label1, label2, is_certain,
               positive_review AS pos, negative_review AS neg
        FROM movies WHERE movie_id = %s
        """,
        (movie_id,),
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        return None
    row["ott"] = _db_get_ott(movie_id)
    if row["rating"]:
        row["rating"] = float(row["rating"])
    return row


# ─────────────────────────────────────────────────────────────────────────────
# API 엔드포인트
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    mode = "db" if _use_db else "excel"
    return {"status": "ok", "mode": mode, "movie_count": len(_movies)}


@app.get("/api/emotions")
def get_emotions():
    """프론트엔드 버튼 목록용 감정 태그 반환."""
    emotions = [
        "울고 싶어", "웃고 싶어", "감동받고 싶어", "설레고 싶어",
        "긴장감을 느끼고 싶어", "몰입하고 싶어", "반전을 보고 싶어",
        "생각하게 만드는 영화를 보고 싶어", "성장 이야기가 보고 싶어",
        "미장센이 좋은 영화를 보고 싶어", "음악이 좋은 영화를 보고 싶어",
        "가벼운 마음으로 시간을 보내고 싶어", "친구랑 보고 싶어", "가족이랑 보고 싶어",
    ]
    return {"emotions": emotions}


def _get_movies_by_emotion(emotion: str, order_by: str = "rating") -> list[dict]:
    if _use_db:
        return _db_get_movies_by_emotion(emotion, order_by)
    return _excel_get_movies_by_emotion(emotion, order_by)


# 2) @app.get("/api/movies") 엔드포인트를 통째로 이걸로 교체하세요.
@app.get("/api/movies")
def get_movies(emotion: str = "", order_by: str = "rating"):
    """
    감정 기반 영화 추천 목록 반환 (정렬 기능 추가)
    order_by="rating" -> 별점순
    order_by="year"   -> 최신순
    """
    if not emotion:
        raise HTTPException(status_code=400, detail="emotion 파라미터가 필요합니다.")

    # 정렬 기준을 안 넘겨주면 기본값인 "rating"(별점순)으로 동작합니다.
    movies = _get_movies_by_emotion(emotion.strip(), order_by)

    for m in movies:
        m["ott_names"] = [o["platform"] for o in m.get("ott", [])]

    return {"status": "success", "emotion": emotion, "count": len(movies), "data": movies}


@app.get("/api/movies/{movie_id}")
def get_movie_detail(movie_id: str):
    """
    특정 영화의 상세 정보 반환.

    프론트엔드 호출 예시:
      fetch('/api/movies/M001')
    """
    movie = _get_movie_detail(movie_id.upper())
    if not movie:
        raise HTTPException(status_code=404, detail=f"영화 ID '{movie_id}'를 찾을 수 없습니다.")
    movie["ott_names"] = [o["platform"] for o in movie.get("ott", [])]
    return {"status": "success", "data": movie}


@app.get("/api/search")
def search_movies(q: str = ""):
    """제목 키워드 검색."""
    if not q:
        raise HTTPException(status_code=400, detail="q 파라미터가 필요합니다.")

    q_lower = q.lower()
    results = [
        m for m in _movies.values()
        if q_lower in (m.get("title") or "").lower()
    ]
    results.sort(key=lambda x: -(x["rating"] or 0))

    for m in results:
        m["ott_names"] = [o["platform"] for o in m.get("ott", [])]

    return {"status": "success", "query": q, "count": len(results), "data": results}


# ── 정적 파일 서빙 (프론트엔드 HTML) ──────────────────────────────────────────
# index.html 이 같은 폴더에 있으면 / 경로에서 바로 열림
STATIC_DIR = os.path.dirname(os.path.abspath(__file__))

@app.get("/")
def serve_index():
    html_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(html_path):
        return FileResponse(html_path)
    return {"message": "index.html 이 없습니다. 같은 폴더에 넣어주세요."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=5000, reload=True)
